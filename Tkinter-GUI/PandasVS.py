import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, NamedStyle
import sys

# Check if the file path is provided as a command-line argument
if len(sys.argv) != 4:
    print("Usage: python PandasVS.py <input_file_path> <output_file_path> <secondsheet_file_path>")
    sys.exit(1)

# Get the file path from the command-line argument
input_file_path = sys.argv[1]
output_file_path = sys.argv[2]
secondsheet_file_path = sys.argv[3]

# Load the Excel file into a DataFrame
df = pd.read_excel(input_file_path)

# Fill any blanks with 0 for concatenation
df = df.fillna(0)

# Define the specific COT_GROUP value to filter
specific_cot = ['HOSPITAL/HEALTH SYSTEM']

# Filter the DataFrame based on the specific COT_GROUP
df_filtered = df[df['COT_GROUP'].isin(specific_cot)].copy()

df_filtered['MANF_DESC-SKU'] = df_filtered['MANF_DESC'].astype(str) + '-' + df_filtered['SKU'].astype(str)

# List of columns in the order you want them
cols = ['COT_GROUP', 'MANF_DESC', 'SKU','MANF_DESC-SKU', 'PROD_DESC', 'UNSPSC_CODE','UNSPSC_2_DESC', 'UNSPSC_3_DESC','UNSPSC_4_DESC']

# Reorder the columns
df = df_filtered[cols]

# Combine the YEAR and QUARTER columns into a single column named 'year-quarter' with a dash
df_filtered['year-quarter'] = df_filtered['YEAR'].astype(str) + '-Q' + df_filtered['QUARTER'].astype(str)

# Pivot the DataFrame so that 'year-quarter' are the columns and 'REVENUE' and 'UNITS' are the values
pivot_revenue = df_filtered.pivot_table(index=['COT_GROUP', 'MANF_DESC', 'SKU', 'MANF_DESC-SKU', 'PROD_DESC', 'UNSPSC_CODE','UNSPSC_2_DESC', 'UNSPSC_3_DESC','UNSPSC_4_DESC'],
                                        columns='year-quarter',
                                        values='REVENUE',
                                        aggfunc='sum')

pivot_units = df_filtered.pivot_table(index=['COT_GROUP', 'MANF_DESC', 'SKU', 'MANF_DESC-SKU', 'PROD_DESC', 'UNSPSC_CODE','UNSPSC_2_DESC', 'UNSPSC_3_DESC','UNSPSC_4_DESC'],
                                      columns='year-quarter',
                                      values='UNITS',
                                      aggfunc='sum')

# Sort the columns based on the year and quarter values
pivot_revenue = pivot_revenue.reindex(sorted(pivot_revenue.columns), axis=1)
pivot_units = pivot_units.reindex(sorted(pivot_units.columns), axis=1)

# Rename the columns to include 'Revenue' and 'Units'
pivot_revenue.columns = [f"{col} Revenue" for col in pivot_revenue.columns]
pivot_units.columns = [f"{col} Units" for col in pivot_units.columns]

# Interleave the columns of revenue and units
revenue_cols = pivot_revenue.columns.tolist()
units_cols = pivot_units.columns.tolist()

# Ensure both have the same length and correct pairing
interleaved_columns = []
for rev_col, unit_col in zip(revenue_cols, units_cols):
    interleaved_columns.append(rev_col)
    interleaved_columns.append(unit_col)

# Combine the DataFrames
combined = pd.concat([pivot_revenue, pivot_units], axis=1)
combined = combined[interleaved_columns]

# Reset index to flatten the DataFrame
combined.reset_index(inplace=True, drop=False)

# Replace NaN values with blank
#combined = combined.fillna('')

# Copy and modify the original DataFrames
pivot_revenue_copy = pivot_revenue.copy() * 0.98
pivot_units_copy = pivot_units.copy()

# Rename the columns in the copied DataFrames
pivot_revenue_copy.columns = [f'{col} SUM' for col in pivot_revenue_copy.columns]
pivot_units_copy.columns = [f'{col} SUM' for col in pivot_units_copy.columns]

# Create a list of DataFrames in the order you want them interleaved
dfs = [pivot_revenue, pivot_units,  pivot_revenue_copy, pivot_units_copy]

revenue_cols2 = pivot_revenue_copy.columns.tolist()
units_cols2 = pivot_units_copy.columns.tolist()

# Ensure both have the same length and correct pairing
interleaved_columns_2 = []
for rev_col2, unit_col2 in zip(revenue_cols2, units_cols2):
    interleaved_columns_2.append(rev_col2)
    interleaved_columns_2.append(unit_col2)

# Combine the DataFrames
combined = pd.concat([pivot_revenue_copy, pivot_units_copy], axis=1)
combined = combined[interleaved_columns_2]

# Reset index to flatten the DataFrame
combined.reset_index(inplace=True, drop=False)

# Use pd.concat with keys to create a multi-index DataFrame, then swap levels and sort to interleave
combined = pd.concat(dfs, axis=1, keys=range(len(dfs))).swaplevel(0, 1, axis=1).sort_index(axis=1)

# Flatten the column MultiIndex
combined.columns = combined.columns.get_level_values(0)

# Reset index to flatten the DataFrame
combined.reset_index(inplace=True)

# Replace NaN values with blank
#combined = combined.fillna('')

# Interleave the original columns
original_columns = [col for pair in zip(pivot_revenue.columns, pivot_units.columns) for col in pair]

# Interleave the copied columns
copied_columns = [col for pair in zip(pivot_revenue_copy.columns, pivot_units_copy.columns) for col in pair]

combined['REPORTED REVENUE - 2% DIST MARKUP --->'] = ''


# Concatenate the interleaved original and copied columns
columns = cols + original_columns + ['REPORTED REVENUE - 2% DIST MARKUP --->'] + copied_columns 

# Assign the interleaved columns to the combined DataFrame
combined = combined[columns]

# Extract the years from the column names
years = set(col.split('-')[0] for col in pivot_revenue.columns)

# Initialize two dictionaries to separately store the total units and total revenue for each year
units_dict = {year: [] for year in years}
revenue_dict = {year: [] for year in years}

# Iterate over the DataFrame columns once
for col in combined.columns:
    # Split the column name into components
    components = col.split()

    # Check if the column name has the correct format
    if len(components) < 3:
        continue

    year = components[0].split('-')[0]
    type_ = components[1]

    if type_ == "Units":
        units_dict[year].append(col)
    elif type_ == "Revenue":
        revenue_dict[year].append(col)

# Merge the units and revenue dictionaries
totals_dict = {year: {'units': units_dict[year], 'revenue': revenue_dict[year]} for year in years}

# Initialize a new DataFrame to store the total revenue and total units for each year
totals = pd.DataFrame(index=combined.index)

new_column_df2 = pd.DataFrame({'TOTAL ANNUAL REVENUE --->': [None] * len(totals)})

# Concatenate the new column DataFrame with the existing years_data DataFrame
totals = pd.concat([new_column_df2, totals], axis=1)

# First pass: Add all 'TOTAL UNITS' columns for each year
for year in sorted(totals_dict.keys()):
    totals[f'{year} TOTAL UNITS'] = combined[totals_dict[year]['units']].apply(pd.to_numeric, errors='coerce').sum(axis=1)

# Second pass: Add all 'TOTAL REVENUE' columns for each year
for year in sorted(totals_dict.keys()):
    totals[f'{year} TOTAL REVENUE'] = combined[totals_dict[year]['revenue']].apply(pd.to_numeric, errors='coerce').sum(axis=1)

years = set(col.split('-')[0] for col in pivot_revenue.columns)

# Initialize a dictionary to store the total revenue and total units for each year
ASP_dict = {year: {'revenue': [], 'units': []} for year in years}

# Iterate over the DataFrame columns once
for col in combined.columns:
    # Split the column name into components
    components = col.split()

    # Check if the column name has the correct format
    if len(components) < 3:
        continue

    year = components[0].split('-')[0]
    type_ = components[1]

    if type_ in ["Revenue", "Units"]:
        ASP_dict[year][type_.lower()].append(col)

# Initialize a new DataFrame to store the total revenue and total units for each year
ASPs = pd.DataFrame(index=combined.index)

ASPs['ANNUAL ASP --->'] = ''

# For each year, create a new column in the totals DataFrame with the total revenue and total units for that year
for year in sorted(ASP_dict.keys()):
    revenue = combined[ASP_dict[year]['revenue']].apply(pd.to_numeric, errors='coerce').sum(axis=1)
    units = combined[ASP_dict[year]['units']].apply(pd.to_numeric, errors='coerce').sum(axis=1)
    

    ASPs[f'{year} ASP'] = revenue / units

# Extract the quarters from the column names
quarters = set(col.split()[0] for col in pivot_revenue.columns)

# Initialize a dictionary to store the total revenue and total units for each quarter
ASP_6_dict = {quarter: {'revenue': [], 'units': []} for quarter in quarters}

# Iterate over the DataFrame columns once
for col in combined.columns:
    # Split the column name into components
    components = col.split()

    # Check if the column name has the correct format
    if len(components) < 3:
        continue

    quarter = components[0]
    type_ = components[1]

    if type_ in ["Revenue", "Units"]:
        ASP_6_dict[quarter][type_.lower()].append(col)

# Get the last 6 quarters
last_6_quarters = sorted(ASP_6_dict.keys())[-6:]

# Initialize a new DataFrame to store the ASP for the last 6 quarters
ASPs_6_quarters = pd.DataFrame(index=combined.index)

ASPs_6_quarters['QUARTERLY ASP --->'] = ''

# For each of the last 6 quarters, create a new column in the ASP DataFrame with the ASP for that quarter
for quarter in last_6_quarters:
    revenue = combined[ASP_6_dict[quarter]['revenue']].apply(pd.to_numeric, errors='coerce').sum(axis=1)
    units = combined[ASP_6_dict[quarter]['units']].apply(pd.to_numeric, errors='coerce').sum(axis=1)
    ASPs_6_quarters[f'{quarter} ASP'] = revenue / units

# If indices do not match, reset them
df.reset_index(drop=True, inplace=True)
combined.reset_index(drop=True, inplace=True)
totals.reset_index(drop=True, inplace=True)
ASPs.reset_index(drop=True, inplace=True)
ASPs_6_quarters.reset_index(drop=True, inplace=True)

# Concatating files
combined_final = pd.concat([combined, totals, ASPs, ASPs_6_quarters], axis=1, ignore_index=False)

for column in combined.columns:
    if column not in combined_final.columns and column != 'index':
        combined_final[column] = np.nan  

# Save to Excel without styling
#combined_final.to_excel('.tempfile.xlsx', index=False)

secondsheetdf = pd.read_excel(secondsheet_file_path)

years_data = pd.DataFrame()

for column in secondsheetdf.columns:
    column_name = str(column)
    if column_name.isdigit() or '%' in column_name or '20' in column_name or '21' in column_name:
        years_data[column_name] = secondsheetdf[column]

# Filter df2 to only include rows with matching 'manf-sku' values
secondsheetdf_filtered = secondsheetdf[secondsheetdf['MANF_DESC-SKU'].isin(combined_final['MANF_DESC-SKU'])]

# Ensure combined_cols has 'MANF_DESC-SKU' for matching
combined_cols = pd.concat([combined, ASPs, ASPs_6_quarters, totals], axis=1)

# Filter combined_cols to only include rows with matching 'MANF_DESC-SKU' values from filtered secondsheetdf
combined_filtered = combined_cols[combined_cols['MANF_DESC-SKU'].isin(secondsheetdf_filtered['MANF_DESC-SKU'])]

# Keep only the specified columns from the filtered secondsheetdf
secondsheetcols = [ 'MANF_DESC-SKU', 'Market Segment', 'New/Blend/Repr', 'Supplier', 'If Blend/Repr: OEM Brand', 'Matt Length Nominal HTI', 'Matt Width Nominal HTI', 'Includes Chux', 'Use for Price']
secondsheetdf = secondsheetdf_filtered[secondsheetcols]

# Merge on MANF_DESC-SKU
merged_df = pd.merge(secondsheetdf, combined_filtered, on='MANF_DESC-SKU', how='inner')

merged_df.reset_index(drop=True, inplace=True)
years_data.reset_index(drop=True, inplace=True)

new_column_df = pd.DataFrame({'% THROUGH DISTRIBUTIONS --->': [None] * len(years_data)})

# Concatenate the new column DataFrame with the existing years_data DataFrame
years_data = pd.concat([new_column_df, years_data], axis=1)
merged_df = pd.concat([merged_df, years_data], axis=1)

totals = totals.sort_index()
years_data = years_data.sort_index()

# Identify columns in merged_df that contain the word "TOTAL"
total_columns = [col for col in merged_df.columns if 'TOTAL' in col]
# Create a new DataFrame with these columns
total_columns_df = merged_df[total_columns]

years_data.reset_index(drop=True, inplace=True)

units_columns = [col for col in totals.columns if 'TOTAL UNITS' in col] 

# Initialize a new DataFrame to store the results
units_comp_df = pd.DataFrame(index=merged_df.index)

# Loop through each column in the units_columns
for col in units_columns:
    # Extract the year from the column name
    year = col.split()[0]
    
    # Check if there is a corresponding column in years_data with percentage values
    if year in years_data.columns:
        total_columns_df.index = merged_df.index
        # Perform the division and store the result in units_comp_df
        units_comp_df[col] = total_columns_df[col].div(years_data[year], axis=0)  # Convert to percentage
        
# Add column headers 
units_comp_df.columns = [f'Extrapolated {col}' for col in units_comp_df.columns]

# Reset index if necessary
units_comp_df.reset_index(drop=True, inplace=True)

# Merge units_comp_df with the final DataFrame
merged_df = pd.concat([merged_df, units_comp_df], axis=1)

# Sort total columns for the revenue columns
revenue_columns = [col for col in totals.columns if 'TOTAL REVENUE' in col] #?

# Initialize a new DataFrame to store the results
revenue_comp_df = pd.DataFrame(index=merged_df.index)#?

# Loop through each column in the revenue_columns
for col in revenue_columns:#good
    # Extract the year from the column name
    year = col.split()[0]#good

    # Check if there is a corresponding column in years_data with percentage values
    if year in years_data.columns:
        total_columns_df = total_columns_df.sort_index()
        years_data = years_data.sort_index()
        # Perform the division and store the result in revenue_comp_df
        revenue_comp_df[col] = total_columns_df[col].div(years_data[year], axis=0) 

revenue_comp_df.columns = [f'Extrapolated {col}' for col in revenue_comp_df.columns]

# Reset index if necessary
revenue_comp_df.reset_index(drop=True, inplace=True)

# Merge percentages_df with the final DataFrame
merged_df = pd.concat([merged_df, revenue_comp_df], axis=1, ignore_index=False)

# Save data to an Excel file with two sheets
with pd.ExcelWriter('.tempfile.xlsx', engine='openpyxl') as writer:
    combined_final.to_excel(writer, sheet_name='Sheet1', index=False)  # Write the first sheet
    merged_df.to_excel(writer, sheet_name='Sheet2', index=False)  # Write the second sheet

# Load the workbook
workbook = openpyxl.load_workbook('.tempfile.xlsx')

units_format = '#,##0'
revenue_format = '"$"#,##0'
asp_format = '"$"#,##0.00'

units_style = NamedStyle(name="units_style", number_format=units_format)
revenue_style = NamedStyle(name="revenue_style", number_format=revenue_format)
asp_style = NamedStyle(name="asp_style", number_format=asp_format)

# Register the styles with the workbook
workbook.add_named_style(units_style)
workbook.add_named_style(revenue_style)
workbook.add_named_style(asp_style)


# Define the fills
col1_fill = PatternFill(start_color='a1cc35', end_color='a1cc35', fill_type='solid')
col2_fill = PatternFill(start_color='FFC0CB', end_color='FFC0CB', fill_type='solid')
col3_fill = PatternFill(start_color='eaf24b', end_color='eaf24b', fill_type='solid')
col4_fill = PatternFill(start_color='f56936', end_color='f56936', fill_type='solid')
col5_fill = PatternFill(start_color='6e93fa', end_color='6e93fa', fill_type='solid')
col6_fill = PatternFill(start_color='b491ff', end_color='b491ff', fill_type='solid')
col7_fill = PatternFill(start_color='03fc94', end_color='03fc94', fill_type='solid')
col8_fill = PatternFill(start_color='03fcd7', end_color='03fcd7', fill_type='solid')
col9_fill = PatternFill(start_color='ba03fc', end_color='ba03fc', fill_type='solid')
col10_fill = PatternFill(start_color='03d3fc', end_color='03d3fc', fill_type='solid')

# Combine lists for columns to fill
columns_to_fill1 = cols
columns_to_fill2 = interleaved_columns
columns_to_fill3 = interleaved_columns_2 + ['REPORTED REVENUE - 2% DIST MARKUP --->']
columns_to_fill4 = totals
columns_to_fill5 = ['TOTAL ANNUAL REVENUE --->']
columns_to_fill6 = ASPs
columns_to_fill7 = ASPs_6_quarters
columns_to_fill8 = ['% THROUGH DISTRIBUTIONS --->']
columns_to_fill9 = secondsheetcols
columns_to_fill10 = years_data
columns_to_fill11 = units_comp_df
columns_to_fill12 = revenue_comp_df

# Function to apply fills to a given sheet
def apply_fills(sheet):
    for col in sheet.iter_cols(min_row=1, max_row=1, values_only=False):
        if col[0].value in columns_to_fill1:
            for cell in col:
                cell.fill = col1_fill
        if col[0].value in columns_to_fill2:
            for cell in col:
                cell.fill = col2_fill
        if col[0].value in columns_to_fill3:
            for cell in col:
                cell.fill = col3_fill
        if col[0].value in columns_to_fill4:
            for cell in col:
                cell.fill = col4_fill
        if col[0].value in columns_to_fill5:
            for cell in col:
                cell.fill = col4_fill
        if col[0].value in columns_to_fill6:
            for cell in col:
                cell.fill = col5_fill
        if col[0].value in columns_to_fill7:
            for cell in col:
                cell.fill = col6_fill
        if col[0].value in columns_to_fill8:
            for cell in col:
                cell.fill = col7_fill
        if col[0].value in columns_to_fill9:
            for cell in col:
                cell.fill = col8_fill
        if col[0].value in columns_to_fill10:
            for cell in col:
                cell.fill = col7_fill
        if col[0].value in columns_to_fill11:
            for cell in col:
                cell.fill = col9_fill
        if col[0].value in columns_to_fill12:
            for cell in col:
                cell.fill = col10_fill

def apply_format(sheet):
    for col in sheet.iter_cols(min_row=1, max_row=sheet.max_row, values_only=False):
        header = col[0].value 
    
        if 'units' in header.lower():
            style = units_style
            
        elif 'revenue' in header.lower():
            style = revenue_style
        elif 'asp' in header.lower():
            style = asp_style
        else:
            continue

        # Apply style and fill to all cells in the column (excluding the header)
        for cell in col[1:]:
            cell.style = style



# Apply fills to both sheets
apply_fills(workbook['Sheet1'])
apply_fills(workbook['Sheet2'])

apply_format(workbook['Sheet1'])
apply_format(workbook['Sheet2'])

# Save the workbook
workbook.save(output_file_path)